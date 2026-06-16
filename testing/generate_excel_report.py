import os
import time
import sys
import glob
import hashlib
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def get_file_hash(filepath, method="md5"):
    hasher = hashlib.new(method)
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hasher.update(chunk)
    return hasher.hexdigest()

def read_result_file(filepath):
    results = {"STATUS": "N/A", "MESSAGE": "No test executed."}
    if os.path.exists(filepath):
        with open(filepath, "r") as f:
            for line in f:
                if "=" in line:
                    key, value = line.strip().split("=", 1)
                    results[key] = value
    return results

def parse_trivy_json(filepath):
    summary = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
    vulnerabilities = []
    if os.path.exists(filepath):
        try:
            with open(filepath, "r") as f:
                data = json.load(f)
            if "Results" in data:
                for result in data["Results"]:
                    target = result.get("Target", "Unknown")
                    for vuln in result.get("Vulnerabilities", []):
                        severity = vuln.get("Severity", "UNKNOWN").upper()
                        if severity in summary:
                            summary[severity] += 1
                        vulnerabilities.append({
                            "id": vuln.get("VulnerabilityID", "N/A"),
                            "package": vuln.get("PkgName", "N/A"),
                            "installed": vuln.get("InstalledVersion", "N/A"),
                            "fixed": vuln.get("FixedVersion", "N/A"),
                            "severity": severity,
                            "target": target,
                            "title": vuln.get("Title", "N/A")
                        })
        except Exception as e:
            print(f"Error parsing Trivy JSON: {e}")
    return summary, vulnerabilities

def generate_excel():
    print("Generating Excel report...")
    
    # 1. Gather APK Details
    apk_dir = "frontend_temp/build/app/outputs/flutter-apk"
    apk_pattern = os.path.join(apk_dir, "*release.apk")
    apks = glob.glob(apk_pattern)
    
    apk_info = {"name": "N/A", "size": "N/A", "md5": "N/A"}
    if apks:
        apk_path = apks[0]
        # Prefer x86_64 apk for the emulator if available
        for apk in apks:
            if "x86_64" in apk:
                apk_path = apk
                break
        apk_info["name"] = os.path.basename(apk_path)
        apk_info["size"] = f"{os.path.getsize(apk_path) / (1024 * 1024):.2f} MB"
        apk_info["md5"] = get_file_hash(apk_path)
        
    # 2. Gather Test Results
    appium_res = read_result_file("testing/appium_result.txt")
    selenium_res = read_result_file("testing/selenium_result.txt")
    
    # 3. Gather Vulnerability Scans
    trivy_summary, trivy_list = parse_trivy_json("testing/trivy-results.json")
    
    # 4. Create Workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Dashboard
    ws1 = wb.active
    ws1.title = "QA & Build Dashboard"
    ws1.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    sub_header_font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    regular_font = Font(name="Segoe UI", size=10)
    
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # soft green
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # soft red
    neutral_fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Title Block
    ws1.merge_cells("A1:D1")
    title_cell = ws1["A1"]
    title_cell.value = "NutriSync Project Build & QA Report"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40
    
    # Section 1: APK Build Details
    ws1["A3"] = "1. APK BUILD INFO"
    ws1["A3"].font = sub_header_font
    
    apk_headers = ["Property", "Value"]
    for col_idx, h in enumerate(apk_headers, start=1):
        cell = ws1.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = navy_fill
        cell.border = thin_border
        
    apk_data = [
        ("File Name", apk_info["name"]),
        ("File Size", apk_info["size"]),
        ("MD5 Checksum", apk_info["md5"]),
        ("Build Time", time.strftime("%Y-%m-%d %H:%M:%S UTC"))
    ]
    
    for row_idx, (prop, val) in enumerate(apk_data, start=5):
        c1 = ws1.cell(row=row_idx, column=1, value=prop)
        c2 = ws1.cell(row=row_idx, column=2, value=val)
        c1.font = bold_font
        c2.font = regular_font
        c1.border = thin_border
        c2.border = thin_border
        c1.fill = light_blue_fill
        
    # Section 2: Automated UI Test Status
    ws1["A11"] = "2. AUTOMATED UI TESTING"
    ws1["A11"].font = sub_header_font
    
    test_headers = ["Test Suite", "Status", "Message"]
    for col_idx, h in enumerate(test_headers, start=1):
        cell = ws1.cell(row=12, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = navy_fill
        cell.border = thin_border
        
    test_data = [
        ("Mobile Automation (Appium)", appium_res["STATUS"], appium_res["MESSAGE"]),
        ("Web Automation (Selenium)", selenium_res["STATUS"], selenium_res["MESSAGE"])
    ]
    
    for row_idx, (suite, status, msg) in enumerate(test_data, start=13):
        c1 = ws1.cell(row=row_idx, column=1, value=suite)
        c2 = ws1.cell(row=row_idx, column=2, value=status)
        c3 = ws1.cell(row=row_idx, column=3, value=msg)
        
        c1.font = bold_font
        c2.font = bold_font
        c3.font = regular_font
        
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        
        # Color status column
        if status == "PASS":
            c2.fill = pass_fill
        elif status == "FAIL":
            c2.fill = fail_fill
        else:
            c2.fill = neutral_fill
            
    # Section 3: Codebase Vulnerability Scan Summary
    ws1["A17"] = "3. SECURITY VULNERABILITY REPORT"
    ws1["A17"].font = sub_header_font
    
    vuln_headers = ["Severity Level", "Vulnerability Count"]
    for col_idx, h in enumerate(vuln_headers, start=1):
        cell = ws1.cell(row=18, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = navy_fill
        cell.border = thin_border
        
    vuln_data = [
        ("CRITICAL", trivy_summary["CRITICAL"]),
        ("HIGH", trivy_summary["HIGH"]),
        ("MEDIUM", trivy_summary["MEDIUM"]),
        ("LOW", trivy_summary["LOW"])
    ]
    
    for row_idx, (sev, count) in enumerate(vuln_data, start=19):
        c1 = ws1.cell(row=row_idx, column=1, value=sev)
        c2 = ws1.cell(row=row_idx, column=2, value=count)
        
        c1.font = bold_font
        c2.font = regular_font
        
        c1.border = thin_border
        c2.border = thin_border
        
        # Soft alert fills for severity
        if sev == "CRITICAL" and count > 0:
            c1.fill = fail_fill
        elif sev == "HIGH" and count > 0:
            c1.fill = fail_fill
        else:
            c1.fill = light_blue_fill
            
    # Auto-fit columns for Dashboard
    for col in ws1.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.coordinate in ws1.merged_cells:
                continue
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    # Sheet 2: Vulnerability Details (if any exist)
    if trivy_list:
        ws2 = wb.create_sheet(title="Security Vulnerability Details")
        ws2.views.sheetView[0].showGridLines = True
        
        detail_headers = ["ID", "Target File", "Package", "Installed", "Fixed", "Severity", "Description"]
        for col_idx, h in enumerate(detail_headers, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.border = thin_border
            
        for row_idx, v in enumerate(trivy_list, start=2):
            ws2.cell(row=row_idx, column=1, value=v["id"]).font = bold_font
            ws2.cell(row=row_idx, column=2, value=v["target"]).font = regular_font
            ws2.cell(row=row_idx, column=3, value=v["package"]).font = regular_font
            ws2.cell(row=row_idx, column=4, value=v["installed"]).font = regular_font
            ws2.cell(row=row_idx, column=5, value=v["fixed"]).font = regular_font
            
            c_sev = ws2.cell(row=row_idx, column=6, value=v["severity"])
            c_sev.font = bold_font
            if v["severity"] in ("CRITICAL", "HIGH"):
                c_sev.fill = fail_fill
            else:
                c_sev.fill = light_blue_fill
                
            ws2.cell(row=row_idx, column=7, value=v["title"]).font = regular_font
            
            # Apply borders
            for c in range(1, 8):
                ws2.cell(row=row_idx, column=c).border = thin_border
                
        # Auto-fit columns for sheet 2
        for col in ws2.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40) # cap width at 40
            
    # Save Workbook
    output_path = "testing/build_report.xlsx"
    wb.save(output_path)
    print(f"Report successfully saved to {output_path}")

if __name__ == "__main__":
    generate_excel()
